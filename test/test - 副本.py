import pymssql
import openpyxl
import pymysql
import datetime
from win32com.client import DispatchEx

serverIP = '134.96.146.25'
userName = 'Jhdzyw'
passWord = 'Jiaohuan*2003'
dateBase = 'DeviceCheck'
link = pymssql.connect(serverIP, userName, passWord, dateBase, port='1433')
cursor = link.cursor()
print('查询中...')
cursor.execute("select * from jhExc_Summary where iCountTime>'2019-12-2'")
row1 = cursor.fetchone()
d1=[]
while row1:
    d1.append(row1)
    row1 = cursor.fetchone()
cursor.execute("select * from jhExc_Anomaly where iCountTime>'2019-12-2'")
row2 = cursor.fetchone()
d2=[]
while row2:
    d2.append(row2)
    row2 = cursor.fetchone()
cursor.close()
link.close()
print('查询完成')

wb = openpyxl.load_workbook('模板.xlsx')
ws1 = wb["Sheet1"]
print('正在写入Excel...')
for row1 in range(len(d1)):
    for column1 in range(len(d1[row1])):
        ws1.cell(row=row1+2, column=column1+1, value=d1[row1][column1])
ws2 = wb["Sheet2"]
for row2 in range(len(d2)):
    for column2 in range(len(d2[row2])):
        ws2.cell(row=row2+2, column=column2+1, value=d2[row2][column2])
wb.save('example.xlsx')
print('文件保存成功')

xlApp = DispatchEx("Excel.Application")
xlApp.Visible = False
xlBook = xlApp.Workbooks.Open('b.xlsx')
xlBook.Save()
xlBook.Close()
print('刷新成功')



#host='172.20.8.146'
#user='root'
#password='Rot@2017'
#database='dzywdb'
#conn = pymysql.connect(host,user,password,database,port=3306)
#cursor = conn.cursor()
#cursor.execute("SELECT MAX(I_ID) FROM T_MAINTENANCE_AGG")
#row = cursor.fetchone()
#while row:
##    print(row)
#    row = cursor.fetchone()
#return row
#conn.commit()
#cursor.close()
#conn.close()