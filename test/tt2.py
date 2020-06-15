import datetime
import openpyxl
from win32com.client import DispatchEx

a1=(178, 6, 82186, 82183, 81983, 200, 3, 1, datetime.datetime(2019, 12, 2, 0, 33, 1, 177000))
a2=(179, 6, 82271, 82270, 82069, 201, 1, 1, datetime.datetime(2019, 12, 3, 0, 33, 1, 30000))
a3=(180, 6, 82255, 82255, 82056, 199, 0, 1, datetime.datetime(2019, 12, 4, 0, 33, 1, 97000))
d=[]
i=1
while i<4:
    d.append(eval('a'+str(i)))
    i+=1
wb = openpyxl.load_workbook('模板.xlsx')
ws = wb["Sheet1"]
for row in range(len(d)):
    for column in range(len(d[row])):
        ws.cell(row=row+2, column=column+1, value=d[row][column])
    ws['A'+str(row+2)] = 400+row+1
tm = datetime.datetime.now().strftime("%Y%m%d-%H_%M_%S")
filename = tm+'.xlsx'
absolute_path = 'E:\Freely_Lzk\\'+filename
wb.save(filename)

print('文件刷新中...')
xlApp = DispatchEx("Excel.Application")
xlApp.Visible = False
xlBook = xlApp.Workbooks.Open(absolute_path)
xlBook.Save()
xlBook.Close()
print('刷新成功')

wb = openpyxl.load_workbook(filename,data_only = True)
ws2 = wb["Sheet1"]
sql=''
for a in range(len(d)):
    sql+=ws2['J'+str(a+2)].value+'\n'
print(sql)
print(type(sql))

#my_file=open('my file.txt','a')   # 'a'=append 以增加内容的形式打开
#my_file.write(sql\n)
#my_file.close()