import pymssql
import openpyxl
import pymysql
import datetime
import os,sys
import re
from win32com.client import DispatchEx
#project_path = os.path.abspath(os.path.dirname(os.path.dirname(__file__)))
project_path = os.path.abspath(os.path.dirname(__file__))
sys.path.append(project_path)
from SQL import test_sql1,test_sql2
template = os.path.join(project_path,"同步模板.xlsx")
sqlpath = os.path.join(project_path,"SQL.py")
tm = datetime.datetime.now().strftime("%Y%m%d-%H_%M_%S")
dt = datetime.datetime.now().strftime("%Y-%m-%d")
filename = '同步-'+tm+'.xlsx'
savepath = os.path.join(project_path,filename)

f = open(sqlpath,'r',encoding='utf-8')
dt = datetime.datetime.now().strftime("%Y-%m-%d")
for line in f.readlines():
    if 'test_sql1' in line:
        t = re.findall(r"(\d{4}-\d{1,2}-\d{1,2})", line)
        if t[0] == dt:
            print ('jhExc_Summary表查询时间为本日，检查sqlpath配置')
            sys.exit()
        else:
            pass
    elif 'test_sql2' in line:
        t = re.findall(r"(\d{4}-\d{1,2}-\d{1,2})", line)
        if t[0] == dt:
            print ('jhExc_Anomaly表查询时间为本日，检查sqlpath配置')
            sys.exit()
        else:
            pass
f.close()

print('查询中...')
serverIP = '134.96.146.25'
userName = 'Jhdzyw'
passWord = 'Jiaohuan*2003'
dateBase = 'DeviceCheck'
link = pymssql.connect(serverIP, userName, passWord, dateBase, port='1433')
cursor = link.cursor()
cursor.execute(test_sql1)
row1 = cursor.fetchone()
d1=[]
while row1:
    d1.append(row1)
    row1 = cursor.fetchone()
cursor.execute(test_sql2)
row2 = cursor.fetchone()
d2=[]
while row2:
    d2.append(row2)
    row2 = cursor.fetchone()
cursor.close()
link.close()

host='172.20.8.146'
user='root'
password='Rot@2017'
database='dzywdb'
conn = pymysql.connect(host,user,password,database,port=3306)
cursor = conn.cursor()
cursor.execute("SELECT MAX(I_ID) FROM T_MAINTENANCE_AGG")
row = cursor.fetchone()
while row:
    max_agg = row[0]
    row = cursor.fetchone()
max_agg = int(max_agg)
cursor.execute("SELECT MAX(I_ID) FROM T_MAINTENANCE_ANOMALY")
row = cursor.fetchone()
while row:
    max_anomaly = row[0]
    row = cursor.fetchone()
max_anomaly = int(max_anomaly)
cursor.close()
conn.close()
print('查询完成')

wb = openpyxl.load_workbook(template)
ws1 = wb["Sheet1"]
print('正在写入Excel...')
for row1 in range(len(d1)):
    for column1 in range(len(d1[row1])):
        ws1.cell(row=row1+2, column=column1+1, value=d1[row1][column1])
    ws1['A'+str(row1+2)] = max_agg+1+row1
ws2 = wb["Sheet2"]
for row2 in range(len(d2)):
    for column2 in range(len(d2[row2])):
        ws2.cell(row=row2+2, column=column2+1, value=d2[row2][column2])
    ws2['A'+str(row2+2)] = max_anomaly+1+row2
wb.save(savepath)
wb.close()
print('文件保存成功，路径： '+savepath)

print('文件刷新中...')
xlApp = DispatchEx("Excel.Application")
xlApp.Visible = False
xlBook = xlApp.Workbooks.Open(savepath)
xlBook.Save()
xlBook.Close()
xlApp.Quit()
print('刷新成功')

print('SQL写入中...')
sql1 = ""
sql2 = ""
conn = pymysql.connect(host,user,password,database,port=3306,charset='utf8')
cursor = conn.cursor()
wb1 = openpyxl.load_workbook(savepath,data_only = True)
ws3 = wb1["Sheet1"]
for a in range(len(d1)):
    sql1=ws3['J'+str(a+2)].value
    cursor.execute(sql1)
ws4 = wb1["Sheet2"]
for b in range(len(d2)):
    sql2=ws4['J'+str(b+2)].value
    cursor.execute(sql2)
wb1.close()
conn.commit()
cursor.close()
conn.close()
f = open(sqlpath,'r',encoding='utf-8')
f2 = open(sqlpath,'r+',encoding='utf-8')
dt = datetime.datetime.now().strftime("%Y-%m-%d")
for line in f.readlines():
    if 'test_sql1' in line:
        t = re.findall(r"(\d{4}-\d{1,2}-\d{1,2})", line)
        line = line.replace(t[0],dt)
    elif 'test_sql2' in line:
        t = re.findall(r"(\d{4}-\d{1,2}-\d{1,2})", line)
        line = line.replace(t[0],dt)
    f2.write(line)
f.close()
f2.close()
print('ALL DONE')
sys.exit()
